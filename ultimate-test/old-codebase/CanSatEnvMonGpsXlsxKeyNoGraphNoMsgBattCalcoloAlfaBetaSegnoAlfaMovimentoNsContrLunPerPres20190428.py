import os
import sys
import serial    
import time
import math
import openpyxl

from shutil import copyfile

from datetime import datetime
 
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import matplotlib.animation as animation
 
from matplotlib import gridspec



ns = serial.Serial('COM13', 9600, timeout = 0)



x = datetime.now()
xstr = str(datetime.now())

# 2019-04-07 21:38:36.959884

xyear = xstr[0:4]
xmonth = xstr[5:7]
xday = xstr[8:10]
xhours = xstr[11:13]
xminutes = xstr[14:16]
xseconds = xstr[17:19]

out_fileName = 'dataFromNs' + xyear + xmonth + xday + xhours + xminutes + xseconds + ".txt"

out_file = open(out_fileName,'w')   # per memorizzare comandi e risposte a/da NexStar
out_file.close()


LogFileName = 'LogFileEnvAccGps' + xyear + xmonth + xday + xhours + xminutes + xseconds + ".xlsx"
# File = 'LogFileEnvAndGps.xlsx'
# envLogFileName = xyear + xmonth + xday + xhours + xminutes + xseconds + "Env.txt"

#out_file = open(LogFileName,'w')   #non servono più, usiamo excel
#out_file.close()

#envout_file = open(envLogFileName,'w')     #non servono più, usiamo excel
#envout_file.close() 

# os.rename('a.txt', 'b.kml')

copyfile('LogFileEnvAccGpsStandard.xlsx', LogFileName)

 
arduino = serial.Serial('COM14', 115200, timeout = 0)   # per Mega su VivoBook
time.sleep(1)       # era 2



     
xArray = []         # contiene la coordinata x, tempo
hArray = []         # array dell'umidità
pArray = []         #array della pressione
aArray = []       #array dell'altezza
tArray = []       #array della temperaturaq
axArray = []       #array dell'accelerazione lungo x
ayArray = []       #array dell'accelerazione lungo y
azArray = []       #array dell'accelerazione lungo z


def process_key(event):
    print("Key:", event.key)
    if event.key == 'q':
        print("Chiusura programma")
        quit()


# stringa env tipo, compreso il #

    #S13:14H26.43P97528.05A320.51T22.22Ax0.13Ay-9.84Az0.82B7.53E*

def process_timer():

    global contRow

    pressione = 0.0
    altezza = 0.0
    temperatura = 0.0

    try:        
        #print("Leggo eventuali arrivi")    
        #for i in range(1,4):               # inserito per ripulire i buffer di arrivo, ma ora non serve
        #    inputLine = arduino.readline() # ricevo echo locale comando inviato
        #    print(inputLine)
        # print("Invio comando fasullo per ric. sensori")
        # arduino.write(b'j')
        inputLine = arduino.readline() 
        print(inputLine)        

        # out_file = open(LogFileName,'a')        #spostato sotto
        # envout_file = open(envLogFileName,'a')    # spostato sotto

        out_file = open(out_fileName,'a')
        
        
        inputLineStr = str(inputLine)   #necessario per la successiva find
        if len(inputLineStr ) < 20:     #era 4
            pass
        elif inputLineStr.find("#")== 2:
            #print("Eureka #")
                
            if inputLineStr.find("S")== 3:
                #print("Eureka S")
                if inputLineStr.find("E*") > 0:
                    #print("Stringa iniziata e terminata correttamente")
                    # print(inputLineStr)
                    #arduino.flushInput()
                    HPos = inputLineStr.find("H", 0)        # *** VERIFICARE I VALORI NULLI DELLA POSIZ:???
                    #print(HPos)
                    PPos = inputLineStr.find("P", HPos + 1)
                    #print(PPos)
                    APos = inputLineStr.find("A", PPos + 1)
                    #print(APos)
                    TPos = inputLineStr.find("T", APos + 1)
                    #print(TPos)
                    AxPos = inputLineStr.find("Ax", TPos + 1)
                    #print(AxPos)
                    AyPos = inputLineStr.find("Ay", AxPos + 1)
                    #print(AyPos)
                    AzPos = inputLineStr.find("Az", AyPos + 1)
                    #print(AzPos)
                    BPos = inputLineStr.find("B", AzPos + 1)
                    #print(BPos)
                    EPos = inputLineStr.find("E", BPos + 1)
                    #print(EPos)
                        
                    timer = inputLineStr[4:HPos]
                    #print(timer)
                    umidita= inputLineStr[HPos+1:PPos]
                    #print(umidita)
                    pressione = inputLineStr[PPos+1:APos]
                    #print(pressione)
                    altezza = inputLineStr[APos+1:TPos]
                    #print(altezza)
                    temperatura = inputLineStr[TPos+1:AxPos]
                    #print(temperatura)
                    accelX = inputLineStr[AxPos+2:AyPos]
                    #print(accelX)
                    accelY = inputLineStr[AyPos+2:AzPos]    
                    #print(accelY)   
                    accelZ = inputLineStr[AzPos+2:BPos]
                    #print(accelZ)
                    vBatt = inputLineStr[BPos+1:EPos]
                    #if float(vBatt) < 7.55:
                        #print("ATTENZIONE!!!! tensione di batteria critica: ", end = '')
                        #print(vBatt)
     
                    x = datetime.now()
                    xstr = str(datetime.now())

                    # print("Data e ora: " + xstr)
                   
                    xh = int(xstr[11:13])
                    xm = int(xstr[14:16])
                    xs = int(xstr[17:19])
                    totSec = xh*3600+xm*60+xs
                    #print(totSec)

                    if len(xArray) == 20:       #era >=
                        shiftInPlace(xArray, 1)
                        shiftInPlace(hArray, 1)
                        shiftInPlace(pArray, 1)
                        shiftInPlace(aArray, 1)
                        shiftInPlace(tArray, 1)
                        # xArray[19] = totSec
                        xArray[19] = xstr[11:19]
                        hArray[19] = float(umidita)
                        pArray[19] = float(pressione)
                        aArray[19] = float(altezza)
                        tArray[19] = float(temperatura)
                    else:
                        # xArray.append(totSec)
                        xArray.append(xstr[11:19])
                        hArray.append(float(umidita))
                        pArray.append(float(pressione))
                        aArray.append(float(altezza))
                        tArray.append(float(temperatura))

                        #print('Lunghezza vettore x: ')
                        #print(str(len(xArray)))
                        #print('Lunghezza vettore umidità: ')
                        #print(str(len(hArray)))
                        #print('Lunghezza vettore pressione: ')
                        #print(str(len(pArray)))
                        #print('Lunghezza vettore altezza: ')
                        #print(str(len(aArray)))
                        #print('Lunghezza vettore temperatura: ')
                        #print(str(len(tArray)))

                        # print("Tensione batteria: " + str(vBatt))


                    # envout_file = open(envLogFileName,'a')
                    timeStamp = xstr[11:19]
                    # envout_file.write(xstr[11:19])
                    # envout_file.write(";")
                    # envout_file.write(str(pressione))
                    # envout_file.write(";")
                    # envout_file.write(str(altezza))
                    # envout_file.write(";")
                    # envout_file.write(str(temperatura))
                    # envout_file.write(";")
                    # envout_file.write(str(accelX))
                    # envout_file.write(";")
                    # envout_file.write(str(accelY))
                    # envout_file.write(";")
                    # envout_file.write(str(accelZ))
                        
                    # envout_file.write('\n')
                    # envout_file.close()


# calcolo angolo beta

                    disTelRampa = 400.0           # fissata a priori dopo la misura sul campo
                                            #      si potrebbe calcolare dai dati GPS
                

                    wb = openpyxl.load_workbook(filename=LogFileName)
                    ws = wb.get_sheet_by_name('Env')
                    contRow = ws['N1'].value
                    contRowPrec = contRow - 1

                    hPrec = ws.cell(contRowPrec, 4).value
                    betaGradTotalePrec = ws.cell(contRowPrec, 7).value

                    print("Altezza precedente: ", end = "")
                    print(str(hPrec))

                    hNuovo = float(altezza)
                    # hNuovo = float(ws.cell(contRow, 4).value)

                    print("Altezza nuova: ", end = "")
                    print(str(hNuovo))

                    tgTetaPrec = hPrec/disTelRampa
                    tetaPrec = math.atan(tgTetaPrec)
                    tetaPrecGrad = (tetaPrec * 180)/math.pi

                    print("Teta precedente: ", end = "")
                    print(str(tetaPrecGrad))

                    tgTetaNuovo = hNuovo/disTelRampa
                    tetaNuovo = math.atan(tgTetaNuovo)
                    tetaNuovoGrad = (tetaNuovo * 180)/math.pi

                    print("Teta nuovo: ", end = "")
                    print(str(tetaNuovoGrad))

                    betaGrad = tetaPrecGrad - tetaNuovoGrad       # beta > 0 cansat scende, altrimenti sale

                    #if contRow == 3:                   # al primo rilevamento non ho un dato precedente utile
                     #   betaGrad = 60.0               # valore iniziale scelto per comodità di osservazione

                    print("Angolo beta: ", end = "")
                    print(betaGrad)

                    wb = openpyxl.load_workbook(filename=LogFileName)
                    ws = wb.get_sheet_by_name('Env')
                    contRow = ws['N1'].value
                    contRowPrec = contRow - 1

                    betaGradTotalePrec = ws.cell(contRowPrec, 7).value
                     
                    betaGradTotale = betaGradTotalePrec + betaGrad


# MOVIMENTO TELESCOPIO PER ALFA e BETA: NB: beta è l'attuale, alfa è quello precedente

#
# Movimento GoTo - N.B.: il movimento GoTo stoppa il movimento slew
#
# posizionamento a ang°
# 360°:65535 = ang°:PosizDec
#
                    if betaGradTotale < 0:
                        betaGradTotale = betaGradTotale + 360
                        
                    angALTTot = betaGradTotale                           #  OCCORRE PRENDERE IL TOTALE DI ALT!!!
                    PosizDecArrALT = int((65535*angALTTot)/360)      #arrotondo per poter convertire in HEX
                    print("PosizDecALT: " + str(PosizDecArrALT))
              
# conversione in Hex
                    PosizHexALT = hex(PosizDecArrALT).lstrip("0x")
                    if len(PosizHexALT) == 0:
                        PosizHexALT = "0001"
                    if len(PosizHexALT) == 2:
                        PosizHexALT = "00" + PosizHexALT
                    if len(PosizHexALT) == 3:
                        PosizHexALT = "0" + PosizHexALT
                    print("PosizHexALT: ", end = "")
                    print(PosizHexALT)

                    wb = openpyxl.load_workbook(filename=LogFileName)
                    ws = wb.get_sheet_by_name('Gps')
                    contRow = ws['N1'].value

                    contRowPrec = (ws['N1'].value) - 1

                    alfaGradTotPrec = ws.cell(contRowPrec, 10).value

                    angAZMTotPrec = alfaGradTotPrec
                    print("AngAZMTotPrec ", end = "")
                    print(angAZMTotPrec)

                    if angAZMTotPrec < 0:
                        angAZMTotPrec = angAZMTotPrec + 360

                    PosizDecArrAZM = int((65535*angAZMTotPrec)/360)      #arrotondo per poter convertire in HEX
                    print("PosizDecAZMPrec: " + str(PosizDecArrAZM))
              
# conversione in Hex
                    PosizHexAZM = hex(PosizDecArrAZM).lstrip("0x")
                    if len(PosizHexAZM) == 0:
                        PosizHexAZM = "0001"
                    if len(PosizHexAZM) == 2:
                        PosizHexAZM = "00" + PosizHexAZM
                    if len(PosizHexAZM) == 3:
                        PosizHexAZM = "0" + PosizHexAZM
                    print("PosizHexAZM: ", end = "")
                    print(PosizHexAZM)

                    
                    comandoPref = "B"                   # 
                    comandoAZM = PosizHexAZM                 #     lascio AZM a 0 NNNNBBBB: qui occorre inserire ang. prec.tot
                    comandoVirg = ","                   #     
                    comandoALT = PosizHexALT                  # posiz. verticale  ALT  

                    comTemp = comandoPref + comandoAZM + comandoVirg + comandoALT
                    comando = bytes(comTemp, 'utf-8')
        
                    print("Invio del comando: " + str(comando))

                    out_file.write("Comando inviato: " + str(comando) + '\n')
                    

                    if len(comando) > 8:
                        ns.write(comando)
                    else:
                        print("COMANDO NON INVIATO!")
                        out_file.write("COMANDO NON INVIATO!")

                    time.sleep(0.1) #un decimo di secondo basta per ricevere la risposta 

                    inputChar = ns.read()               # hashtag
                    print("Risposta ricevuta: ", end = "")
                    print(inputChar)
                    out_file.write("RispostaRicevuta: " + str(inputChar) + '\n')
                    print("----------------------------")

                    time.sleep(1) # era 20  secondi per il posizionamento, ora è continuo e quindi più piccolo

#leggo la posizione attuale
	
                    comando = "Z"                       # Get AZM-ALT   
	
                    print("Invio del comando: " + comando);
	
                    ns.write(b'Z')

                    time.sleep(0.1) #un decimo di secondo basta per ricevere la risposta

                    inputChar = ns.read()               # p.e. “12AB,4000#”
                    response = inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar

                    inputChar = ns.read()               #
                    response = response + inputChar     # hashtag

                    print("Response: " + str(response))

                    print("Calcolo posizione")

                    pos = response[0:9].decode('UTF-8')

                    print("Position: " + str(pos))

                    azmHex = pos[0:4]
                    altHex = pos[5:9]

                    azmInt = int(azmHex, 16)
                    altInt = int(altHex, 16)

                    print("Azimuth in Hex: " + azmHex, end = "")
                    print(" Altitude in Hex: " + altHex)

                    print("Azimuth Int: " + str(azmInt), end = "")
                    print(" Altitude Int: " + str(altInt))

                    print("----------------------------")


# time.sleep(2) #2  secondi per ricevere la risposta


                    # time.sleep(5)               

     
                    wb = openpyxl.load_workbook(filename=LogFileName)
                    ws = wb.get_sheet_by_name('Env')
                    contRow = ws['N1'].value
                    #print("CCCCCCCCCCCCCCCCCC ")
                    #print(contRow)

                    ws.cell(contRow, 1,timeStamp)
                    ws.cell(contRow, 2, float(temperatura))
                    ws.cell(contRow, 3, float(pressione))
                    ws.cell(contRow, 4, float(altezza))
                    ws.cell(contRow, 5, float(vBatt))
                    ws.cell(contRow, 6, betaGrad)
                    ws.cell(contRow, 7, betaGradTotale)
                    
                    contRow = contRow + 1
                    ws['N1'] = contRow
                    
                    ws = wb.get_sheet_by_name('Acc')
                    contRow = ws['N1'].value
                    ws.cell(contRow, 1,timeStamp)
                    ws.cell(contRow, 2, float(accelX))
                    ws.cell(contRow, 3, float(accelY))
                    ws.cell(contRow, 4, float(accelZ))

                            
                    
                    contRow = contRow + 1
                    ws['N1'] = contRow

                    wb.save(LogFileName)



                    

        else:
            # print("Stringa GPS")
            
# b'$GNGGA,232153.000,4506.086309,N,00732.016905,E,1,6,2.18,323.728,M,48.217,M,,*42\n'  #GGA campione per analisi

            print(inputLineStr)  # da commentare se non serve

# inserire qui la verifica del checksum (vedi file CalcoloChecksum...)
            astPos = inputLineStr.find("*", 0)
            cksum = inputLineStr[astPos+1:astPos+3]
            print(cksum)    # commentare dopo verifica algoritmo checksum

            chksumdata = inputLineStr[inputLineStr.find("$")+1:inputLineStr.find("*")]
            print(chksumdata)   # questa è la stringa da verificare

            csum = 0

            for c in chksumdata:
               # XOR'ing value of csum against the next char in line
               # and storing the new XOR value in csum
                csum ^= ord(c)
            print(csum)

            if hex(csum) == hex(int(cksum, 16)):
                print("GGA valida")

                GGAVirPos = inputLineStr.find("GGA,", 0)
                #print(GGAVirPos)
                Vir4Pos = inputLineStr.find(",4", GGAVirPos + 1)    # posizione seconda virgola seguita dal num. 4
                #print(Vir4Pos)                                     #      valido solo per long = 4x!!!
                VirNVirPos = inputLineStr.find(",N,", Vir4Pos + 1)
                #print(VirNVirPos)
                VirEVirPos = inputLineStr.find(",E,", VirNVirPos + 1)   # VERIFICARE CHE I VALORI NON SIANO 0??
                #print(VirEVirPos)
                VirNumPos = inputLineStr.find(",", VirEVirPos + 3)  # posizione virgola prima di num. sat.
                VirDilPos = inputLineStr.find(",", VirNumPos + 1)   # posizione virgola prima di hor. dil.
                VirMAsmPos = inputLineStr.find(",", VirDilPos + 1)  # posizione virgola prima di alt. s. m.
                VirMGeoPos = inputLineStr.find(",M,", VirMAsmPos + 1)   # posizione di ,M, prima di alt. geoid
                VirMCamPos = inputLineStr.find(",M,", VirMGeoPos + 3)   # posizione di ,M, prima di campi vuoti


                oraGPS = inputLineStr[GGAVirPos + 4:Vir4Pos]
                print(oraGPS)
                latGPS = inputLineStr[Vir4Pos + 1:VirNVirPos]
                print(latGPS)
                longGPS = inputLineStr[VirNVirPos + 3:VirEVirPos]
                print(longGPS)
                asmGPS = inputLineStr[VirMAsmPos + 1:VirMGeoPos]
                print(asmGPS)
                ageoGPS = inputLineStr[VirMGeoPos + 3:VirMCamPos]
                print(ageoGPS)

                #out_file = open(LogFileName,'a') 

                #out_file.write(oraGPS)
                #out_file.write(";")
                #out_file.write(latGPS)
                #out_file.write(";")
                #out_file.write(longGPS)
                #out_file.write(";")
                #out_file.write(asmGPS)
                #out_file.write(";")
                #out_file.write(ageoGPS)
                #out_file.write('\n')
                #out_file.close()


                # estrazione gradi, min e decimi di min per il calcolo di gradi e decimi di grado

                latGPSGradi = latGPS[0:2]
                latGPSMinuti = latGPS[2:4]
                latGPSDecimiDiMin = "0." + str(latGPS[5:])

                longGPSGradi = longGPS[0:3]
                longGPSMinuti = longGPS[3:5]
                longGPSDecimiDiMin = "0." + str(longGPS[6:])

                print("             LatitudineGradi: " + latGPSGradi)
                print("           Latitudine Minuti: " + latGPSMinuti)
                print(" Latitudine Decimi di Minuto: " + latGPSDecimiDiMin)
                print("           Longitudine Gradi: " + longGPSGradi)
                print("          Longitudine Minuti: " + longGPSMinuti)
                print("Longitudine Decimi di Minuto: " + longGPSDecimiDiMin)

                # calcolo gradi e decimi di grado

                latGPSGradEDecimi = int(latGPSGradi) + (float(latGPSMinuti)/60) + (float(latGPSDecimiDiMin)/60)
                longGPSGradEDecimi = int(longGPSGradi) + (float(longGPSMinuti)/60) + (float(longGPSDecimiDiMin)/60)

                print(" Lat Gradi e decimi: " + str(latGPSGradEDecimi))
                print("Long Gradi e decimi: " + str(longGPSGradEDecimi))
                
                # conversioneInRadianti

                latGPSRad = latGPSGradEDecimi*3.14159/180
                longGPSRad = longGPSGradEDecimi*3.14159/180
                h = 0   #per ora lo consideriamo trascurabile

                print(" Lat Rad: " + str(latGPSRad))
                print("Long Rad: " + str(longGPSRad))

    # calcolo angolo spostamento nel piano

                # calcolo coordinate XYZ

                raggioEq = 6378137
                raggioPol = 6356752.314
                eccPriQuad = 0.00669437999
                
                coeffN = raggioEq/(math.sqrt(1 - (eccPriQuad)*(math.sin(latGPSRad))**2))
                print("N: " + str(coeffN))
                
                xNuovo = (coeffN + h)*math.cos(latGPSRad)*math.cos(longGPSRad)
                yNuovo = (coeffN + h)*math.cos(latGPSRad)*math.sin(longGPSRad)
                zNuovo = (coeffN - eccPriQuad*coeffN + h)* math.sin(latGPSRad)

                print("xNuovo: " + str(xNuovo))
                print("yNuovo: " + str(yNuovo))
                print("zNuovo: " + str(zNuovo))

                # acquisizione GPS del telescopio e relative conversioni
                
                latTele = 4431.421281    #finestra casa  SETTARE CON DATI MEDICINA RILEVATI DA CANSAT
                # print(latTele)
                longTele = 1138.081523
                # print(longTele)
                
                latTeleGradi = "45"
                latTeleMinuti = "06"
                latTeleDecimiDiMin = "0.07999"

                longTeleGradi = "7"
                longTeleMinuti = "32"
                longTeleDecimiDiMin = "0.003665"
                
                latTeleGradEDecimi = int(latTeleGradi) + (float(latTeleMinuti)/60) + (float(latTeleDecimiDiMin)/60)
                longTeleGradEDecimi = int(longTeleGradi) + (float(longTeleMinuti)/60) + (float(longTeleDecimiDiMin)/60)

                latTeleRad = latTeleGradEDecimi*3.14159/180
                longTeleRad = longTeleGradEDecimi*3.14159/180

                coeffN = raggioEq/(math.sqrt(1 - (eccPriQuad)*(math.sin(latTeleRad))**2))

                xTele = (coeffN + h)*math.cos(latGPSRad)*math.cos(longGPSRad)
                yTele = (coeffN + h)*math.cos(latGPSRad)*math.sin(longGPSRad)
                zTele = (coeffN - eccPriQuad*coeffN + h)* math.sin(latGPSRad)


                # calcolo di L1, L2, L3, cos(alfa) da cui angolo alfa di rotazione

                wb = openpyxl.load_workbook(filename=LogFileName)
                ws = wb.get_sheet_by_name('Gps')
                contRow = ws['N1'].value

                contRowPrec = (ws['N1'].value) - 1

                print(str(contRowPrec))
                
                xPrec = ws.cell(contRowPrec, 6).value
                yPrec = ws.cell(contRowPrec, 7).value
                zPrec = ws.cell(contRowPrec, 8).value

                print(str(xPrec))
                print(str(yPrec))
                print(str(zPrec))

                L1 = math.sqrt((yPrec - yTele)**2 + (xTele - xPrec)**2)
                L2 = math.sqrt((yPrec - yNuovo)**2 + (xNuovo - xPrec)**2)
                L3 = math.sqrt((yNuovo - yTele)**2 + (xNuovo - xTele)**2)

                cosAlfa = ((L3)**2 + (L1)**2 - (L2)**2)/(2*L3*L1)

                alfa = math.acos(cosAlfa)

                alfaGrad = (alfa*180)/math.pi

# calcolo segno di alfa in funzione dell'angolo assoluto (vedi disegno su libretto)

                gammaPrec = math.atan(yPrec/xPrec)
                gammaPrecGrad = (gammaPrec*180)/math.pi
                print("Gamma precedente: ", end = "")
                print(gammaPrecGrad)

                gammaNuovo = math.atan(yNuovo/xNuovo)
                gammaNuovoGrad = (gammaNuovo*180)/math.pi
                print("Gamma nuovo: ", end = "")
                print(gammaNuovoGrad)

                if gammaNuovo < gammaPrec:
                    print("Non deve essere corretto")
                    print(alfaGrad)
                else:
                    print("DEVE ESSERE CORRETTO")
                    alfaGrad = -alfaGrad
                    print(-alfaGrad)
      

                print('Alfa: ', end = "")
                print(str(alfaGrad))

                # time.sleep(5)


# MOVIMENTO TELESCOPIO PER ALFA NOOOOOO

# salvataggio dati su file xlxs


                wb = openpyxl.load_workbook(filename=LogFileName)
                ws = wb.get_sheet_by_name('Gps')
                contRow = ws['N1'].value
                contRowPrec = contRow - 1

                alfaGradTotPrec = ws.cell(contRowPrec, 10).value
                alfaGradTot = alfaGradTotPrec + alfaGrad

                ws.cell(contRow, 1, float(oraGPS))
                ws.cell(contRow, 2, float(latGPS))
                ws.cell(contRow, 3, float(longGPS))
                ws.cell(contRow, 4, float(asmGPS))
                ws.cell(contRow, 5, float(ageoGPS))

                ws.cell(contRow, 6, xNuovo)
                ws.cell(contRow, 7, yNuovo)
                ws.cell(contRow, 8, zNuovo)
                ws.cell(contRow, 9, alfaGrad)
                ws.cell(contRow, 10, alfaGradTot)
                            
                    
                contRow = contRow + 1
                ws['N1'] = contRow

                wb.save(LogFileName)                


                
                
            else:
                print("GGA ERRATA!!!!!")   


            
            ## time.sleep(.5)  #importante per dare tempo all'arrivo del dato
                                # nel caso non fosse un dato vero
                               # tolta per accelerare il flusso
     
    except Exception as e:
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)



 
def animate(i):
    plt.cla   # clear the current axes
    axGraf0.clear()
    axGraf0.set_title('Parametri della discesa')
    axGraf0.set_ylabel('Umidità')
    axGraf0.plot(xArray, hArray)
    axGraf1.clear()
    axGraf1.set_ylabel('Pressione')
    axGraf1.plot(xArray, pArray)
    axGraf2.clear()
    axGraf2.set_ylabel('Altezza')
    axGraf2.plot(xArray, aArray)    
    axGraf3.clear()
    axGraf3.set_ylabel('Temperatura')
    axGraf3.set_xlabel('Ora')
    for label in axGraf3.xaxis.get_ticklabels():
        label.set_rotation(45)
    axGraf3.plot(xArray, tArray)
     
def shift(l, n):
    return l[n:] + l[:n]
 
def shiftInPlace(l, n):
    n = n % len(l)
    head = l[:n]
    l[:n] = []
    l.extend(head)
    return l
 
figTot = plt.figure()
 
axGraf0 = plt.subplot2grid((8,6), (0,0), rowspan = 2, colspan=6)
axGraf1 = plt.subplot2grid((8,6), (2,0), rowspan = 2, colspan=6)
axGraf2 = plt.subplot2grid((8,6), (4,0), rowspan = 2, colspan=6)
axGraf3 = plt.subplot2grid((8,6), (6,0), rowspan = 2, colspan=6)
for label in axGraf3.xaxis.get_ticklabels():
    label.set_rotation(45)

 
figTot.canvas.mpl_connect('key_press_event', process_key)    # futuro: per attivare la funzione process_key
# figTot.canvas.mpl_connect('button_press_event', process_button)

 
timer = figTot.canvas.new_timer(interval=100)  # era 1000, poi era 700, poi 500
timer.add_callback(process_timer)
timer.start()


 
'''
img = plt.imread("Sfondo3DRidMov.png")# per inserire eventuale immagine di sfondo
plt.imshow(img,zorder=0)
'''
 
## ani = animation.FuncAnimation(figTot, animate, interval = 500)    # era 500
plt.show()      # tolta animazione per accelerare il flusso
                # lasciato show per poter terminare con key_press_event
